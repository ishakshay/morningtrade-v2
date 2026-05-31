from dotenv import load_dotenv
load_dotenv()


from flask import Flask, jsonify, request, redirect
from flask_cors import CORS
from flask_sock import Sock
import threading
import json
import time
import math
import os
import cache
from datetime import datetime, timezone, timedelta

from screeners import intraday_booster, nr7, top_movers, momentum_spike, indices, sector_scope
from screeners.base import SYMBOLS, COUNTRY_LABELS
from screeners.market_session import run as get_all_sessions
from screeners.nse_options import fetch_option_chain as fetch_options, get_pcr_history, get_full_chain
from screeners.nse_futures_dashboard import fetch_dashboard as fetch_futures_dashboard
from screeners.nse_market import get_market_overview, get_iv_history, get_vix_history, get_pcr_intraday
from screeners.nse_options import get_strike_pcr_history
from screeners.news_feed import fetch_all_feeds, get_cached_news, get_nse_announcements
from screeners.nse_futures import poll_futures_sentiment, update_latest, get_latest
from screeners.nse_gamma import compute_gamma_blast
from screeners.volume_analytics import compute as compute_volume_analytics
from tv_webhook import register_tv_webhook

EOD_EXPORT_DIR = os.path.join(os.path.dirname(__file__), 'eod_exports')
os.makedirs(EOD_EXPORT_DIR, exist_ok=True)

_eod_exported_today = {'date': None}

def export_eod_spreadsheet():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [eod] openpyxl not installed — run: pip install openpyxl")
        return

    ist = timezone(timedelta(hours=5, minutes=30))
    today = datetime.now(ist).strftime('%Y-%m-%d')
    filename = f'morningtrade_options_{today}.xlsx'
    filepath = os.path.join(EOD_EXPORT_DIR, filename)

    print(f"\n--- EOD Export: {filename} ---")

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font    = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill_ce = PatternFill('solid', start_color='C0392B')   # red for CE
    hdr_fill_pe = PatternFill('solid', start_color='27AE60')   # green for PE
    hdr_fill_nt = PatternFill('solid', start_color='2C3E50')   # dark for neutral
    hdr_fill_vx = PatternFill('solid', start_color='8E44AD')   # purple for VIX
    center      = Alignment(horizontal='center', vertical='center')
    left        = Alignment(horizontal='left',   vertical='center')
    thin        = Side(style='thin', color='DDDDDD')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_font    = Font(name='Arial', size=9)

    def style_header(cell, fill):
        cell.font      = hdr_font
        cell.fill      = fill
        cell.alignment = center
        cell.border    = border

    def style_data(cell, align='center'):
        cell.font      = row_font
        cell.alignment = center if align == 'center' else left
        cell.border    = border

    def write_sheet_header(ws, title, headers, fills):
        ws.append([title])
        ws['A1'].font = Font(name='Arial', bold=True, size=12)
        ws['A1'].fill = PatternFill('solid', start_color='1A252F')
        ws['A1'].font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws.append([])
        ws.append(headers)
        for col_idx, fill in enumerate(fills, 1):
            cell = ws.cell(row=3, column=col_idx)
            style_header(cell, fill)

    # ── Sheet 1: IV History — NIFTY ──────────────────────────────────────────
    for symbol in ['NIFTY', 'BANKNIFTY']:
        ws = wb.active if symbol == 'NIFTY' else wb.create_sheet()
        ws.title = f'IV History {symbol}'

        headers = ['Time', 'Spot Price', 'CE IV %', 'PE IV %', 'Avg IV %', 'Spread (CE-PE)', 'VIX', 'CE/VIX', 'PE/VIX', 'CE LTP', 'PE LTP']
        fills   = [hdr_fill_nt]*2 + [hdr_fill_ce]*2 + [hdr_fill_nt] + [hdr_fill_nt] + [hdr_fill_vx]*3 + [hdr_fill_ce, hdr_fill_pe]
        write_sheet_header(ws, f'IV History — {symbol} — {today}', headers, fills)

        iv_hist = get_iv_history(symbol)
        for row in iv_hist:
            spread = round(row.get('ce_iv', 0) - row.get('pe_iv', 0), 2) if row.get('ce_iv') and row.get('pe_iv') else ''
            vix    = row.get('vix', '')
            ce_vix = round(row.get('ce_iv', 0) / vix, 3) if vix and row.get('ce_iv') else ''
            pe_vix = round(row.get('pe_iv', 0) / vix, 3) if vix and row.get('pe_iv') else ''
            ws.append([
                row.get('time', ''),
                row.get('spot', ''),
                row.get('ce_iv', ''),
                row.get('pe_iv', ''),
                row.get('avg_iv', ''),
                spread,
                vix if vix else '',
                ce_vix,
                pe_vix,
                row.get('ce_ltp', ''),
                row.get('pe_ltp', ''),
            ])

        for r in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for c in r:
                style_data(c)

        col_widths = [8, 12, 8, 8, 8, 12, 8, 8, 8, 8, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: PCR Intraday — NIFTY + BANKNIFTY ────────────────────────────
    for symbol in ['NIFTY', 'BANKNIFTY']:
        ws = wb.create_sheet(f'PCR Intraday {symbol}')
        headers = ['Time', 'PCR (OI)', 'CE COI', 'PE COI', 'COI Diff (PE-CE)', 'Signal']
        fills   = [hdr_fill_nt, hdr_fill_nt, hdr_fill_ce, hdr_fill_pe, hdr_fill_nt, hdr_fill_nt]
        write_sheet_header(ws, f'PCR Intraday — {symbol} — {today}', headers, fills)

        pcr_snaps = get_pcr_intraday(symbol, interval_mins=3)
        for row in reversed(pcr_snaps):  # oldest first
            ws.append([
                row.get('time', ''),
                row.get('pcr', ''),
                row.get('ce_coi', ''),
                row.get('pe_coi', ''),
                row.get('diff', ''),
                row.get('signal', ''),
            ])

        for r in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for c in r:
                style_data(c)

        col_widths = [8, 10, 12, 12, 14, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Strike PCR History — NIFTY + BANKNIFTY ──────────────────────
    for symbol in ['NIFTY', 'BANKNIFTY']:
        ws = wb.create_sheet(f'Strike History {symbol}')

        # Get all strikes from first snapshot
        snaps = get_strike_pcr_history(symbol)
        if snaps:
            all_strikes = sorted(snaps[0].get('strikes', {}).keys(), key=lambda x: int(x), reverse=True)
        else:
            all_strikes = []

        # Headers: Time | Strike1 PCR | Strike1 VolDiff | Strike2 PCR ...
        base_headers = ['Time']
        base_fills   = [hdr_fill_nt]
        for strike in all_strikes:
            base_headers += [f'{strike} PCR', f'{strike} CE COI', f'{strike} PE COI', f'{strike} VolDiff']
            base_fills   += [hdr_fill_nt, hdr_fill_ce, hdr_fill_pe, hdr_fill_nt]

        write_sheet_header(ws, f'Strike PCR History — {symbol} — {today}', base_headers, base_fills)

        for snap in snaps:
            row_data = [snap.get('time', '')]
            for strike in all_strikes:
                entry = snap.get('strikes', {}).get(strike, {})
                if isinstance(entry, dict):
                    row_data += [
                        entry.get('pcr_coi', ''),
                        entry.get('ce_coi', ''),
                        entry.get('pe_coi', ''),
                        entry.get('vol_diff', ''),
                    ]
                else:
                    row_data += ['', '', '', '']
            ws.append(row_data)

        for r in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for c in r:
                style_data(c)

        for i in range(1, len(base_headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 10

    # ── Sheet 4: Strike IV History — NIFTY + BANKNIFTY ───────────────────────
    # Per-strike IV (CE/PE) for ATM ±5, captured every 3 min throughout the session.
    for symbol in ['NIFTY', 'BANKNIFTY']:
        ws = wb.create_sheet(f'Strike IV {symbol}')

        snaps = _strike_iv_history.get(symbol, [])
        if not snaps:
            ws.append([f'No strike IV data captured for {symbol} on {today}'])
            continue

        # Use the LATEST snapshot to determine column order (handles ATM drift mid-session)
        latest_strikes = sorted(int(s) for s in snaps[-1].get('strikes', {}).keys())

        # Headers: Time | Spot | VIX | ATM | <strike> CE IV | <strike> PE IV | <strike> CE LTP | <strike> PE LTP | ...
        base_headers = ['Time', 'Spot', 'VIX', 'ATM']
        base_fills   = [hdr_fill_nt, hdr_fill_nt, hdr_fill_vx, hdr_fill_nt]
        for strike in latest_strikes:
            base_headers += [f'{strike} CE IV', f'{strike} PE IV', f'{strike} CE LTP', f'{strike} PE LTP']
            base_fills   += [hdr_fill_ce, hdr_fill_pe, hdr_fill_ce, hdr_fill_pe]

        write_sheet_header(ws, f'Strike IV History (ATM±5) — {symbol} — {today}', base_headers, base_fills)

        for snap in snaps:
            row_data = [
                snap.get('time', ''),
                snap.get('spot', ''),
                snap.get('vix', ''),
                snap.get('atm', ''),
            ]
            for strike in latest_strikes:
                entry = snap.get('strikes', {}).get(strike) or snap.get('strikes', {}).get(str(strike)) or {}
                row_data += [
                    entry.get('ce_iv', ''),
                    entry.get('pe_iv', ''),
                    entry.get('ce_ltp', ''),
                    entry.get('pe_ltp', ''),
                ]
            ws.append(row_data)

        # Style data rows
        for r in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for c in r:
                style_data(c)

        # Highlight ATM column for each row by tinting the cell background lightly
        atm_highlight = PatternFill('solid', start_color='FFF4D6')
        for r_idx, snap in enumerate(snaps, start=4):
            row_atm = snap.get('atm')
            if not row_atm:
                continue
            try:
                col_offset = latest_strikes.index(int(row_atm))   # 0-based position within strike list
            except ValueError:
                continue
            # Each strike occupies 4 columns starting at col 5 (1-indexed)
            for c_idx in range(5 + col_offset * 4, 5 + col_offset * 4 + 4):
                ws.cell(row=r_idx, column=c_idx).fill = atm_highlight

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 10
        for i in range(5, len(base_headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 10

    # ── Sheet 5: VIX History ──────────────────────────────────────────────────
    ws_vix = wb.create_sheet('VIX History')
    headers = ['Time', 'VIX Value']
    fills   = [hdr_fill_nt, hdr_fill_vx]
    write_sheet_header(ws_vix, f'India VIX Intraday — {today}', headers, fills)

    for row in get_vix_history():
        ws_vix.append([row.get('time', ''), row.get('value', '')])

    for r in ws_vix.iter_rows(min_row=4, max_row=ws_vix.max_row):
        for c in r:
            style_data(c)
    ws_vix.column_dimensions['A'].width = 10
    ws_vix.column_dimensions['B'].width = 12

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(filepath)
    print(f"  [eod] Saved: {filepath}")
    return filepath


def eod_scheduler():
    """Runs in background — triggers export at 3:30 PM IST every trading day."""
    ist = timezone(timedelta(hours=5, minutes=30))
    while True:
        now = datetime.now(ist)
        today_str = now.strftime('%Y-%m-%d')
        # Trigger at 15:30 IST
        if now.hour == 15 and now.minute >= 30 and _eod_exported_today['date'] != today_str:
            # Only on weekdays (Mon-Fri)
            if now.weekday() < 5:
                _eod_exported_today['date'] = today_str
                try:
                    export_eod_spreadsheet()
                except Exception as e:
                    print(f"  [eod] export failed: {e}")
        time.sleep(60)  # check every minute

_news_items_cache = {'data': [], 'ts': 0}
NEWS_CACHE_TTL    = 300  # 5 minutes

def refresh_news():
    while True:
        try:
            print("\n--- Refreshing news ---")
            items    = fetch_all_feeds()
            nse_ann  = get_nse_announcements()
            all_news = (nse_ann + items)[:200]
            global _news_items_cache
            _news_items_cache = {'data': all_news, 'ts': time.time()}
            print(f"  [news] {len(all_news)} items fetched")
        except Exception as e:
            print(f"  [news] refresh error: {e}")
        time.sleep(300)

app = Flask(__name__)
CORS(app, origins=["http://localhost:3000"])
register_tv_webhook(app)

@app.route('/api/news')
def get_news():
    region = request.args.get('region', 'ALL')

    if time.time() - _news_items_cache['ts'] < NEWS_CACHE_TTL and _news_items_cache['data']:
        all_news = list(_news_items_cache['data'])
    else:
        items    = fetch_all_feeds()
        nse      = get_nse_announcements()
        all_news = (nse + items)[:80]

    if region != 'ALL':
        all_news = [n for n in all_news if n.get('region') == region]

    return jsonify(sanitize(all_news))


sock = Sock(app)

SCREENERS      = [intraday_booster, nr7, top_movers, momentum_spike, sector_scope]
COUNTRIES      = list(SYMBOLS.keys())
_options_cache    = {}
_market_cache     = {}
_gamma_cache      = {}
_prev_pcr_3strike = {}
_prev_chain_cache = {}   # holds previous /api/option-chain snapshot per symbol for volume delta computation
_strike_iv_history = {'NIFTY': [], 'BANKNIFTY': []}   # rolling per-strike IV snapshots for ATM±5, capped at 130 per symbol (full session)

def sanitize(obj):
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return 0
        return obj
    if isinstance(obj, dict):
        return {k: sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [sanitize(i) for i in obj]
    return obj
from screeners.mtf_scanner_route import register_scanner_routes
register_scanner_routes(app, sanitize)

def safe_json(data):
    return json.dumps(sanitize(data))

def refresh_futures():
    symbols = ['NIFTY', 'BANKNIFTY']
    while True:
        print("\n--- Refreshing futures sentiment ---")
        for symbol in symbols:
            try:
                payload = poll_futures_sentiment(symbol)
                if payload:
                    update_latest(symbol, payload)
                    print(f"  [futures] {symbol} signal:{payload['signal']} conf:{payload['confidence']} basis:{payload['basis']}")
            except Exception as e:
                print(f"  [futures] {symbol} failed: {e}")
        time.sleep(60)

def refresh_futures_dashboard():
    symbols = ['NIFTY', 'BANKNIFTY']
    while True:
        for symbol in symbols:
            try:
                options_data = _options_cache.get(symbol)
                fetch_futures_dashboard(symbol, options_data)
            except Exception as e:
                print(f"  [futures_dashboard] {symbol} failed: {e}")
        time.sleep(180)

def refresh_market_overview():
    while True:
        print("\n--- Refreshing market overview ---")
        try:
            result = get_market_overview()
            if result:
                _market_cache.update(result)
        except Exception as e:
            print(f"  [market_overview] failed: {e}")
        time.sleep(180)

def refresh_candles():
    """Refresh Nifty/BankNifty OHLC candles every 3 minutes during market hours."""
    from screeners.nse_market import fetch_candles
    while True:
        try:
            nifty_c = fetch_candles('^NSEI')
            bn_c    = fetch_candles('^NSEBANK')
            if nifty_c:
                _market_cache['nifty_candles']    = nifty_c
            if bn_c:
                _market_cache['banknifty_candles'] = bn_c
        except Exception as e:
            print(f"  [candles] failed: {e}")
        time.sleep(180)

def _snapshot_strike_iv(symbol, full_chain_data):
    """
    Capture per-strike IV (CE/PE) for ATM ±5 strikes from current chain data.
    Stores in _strike_iv_history[symbol] for retrieval and EOD export.
    """
    if not full_chain_data:
        return
    chain = full_chain_data.get('full_chain') or full_chain_data.get('chain') or []
    spot  = full_chain_data.get('spot_price', 0)
    atm   = full_chain_data.get('atm_strike', 0)
    if not chain or not atm:
        return

    # Infer strike step (50 NIFTY, 100 BANKNIFTY)
    sorted_strikes = sorted({r.get('strike', 0) for r in chain if r.get('strike')})
    step = 100 if symbol == 'BANKNIFTY' else 50
    if len(sorted_strikes) >= 2:
        diffs = [sorted_strikes[i+1] - sorted_strikes[i] for i in range(len(sorted_strikes)-1)]
        diffs = [d for d in diffs if d > 0]
        if diffs:
            step = min(diffs)

    # Build {strike: {ce_iv, pe_iv, ce_ltp, pe_ltp}} for ATM±5
    strikes_obj = {}
    for k in range(-5, 6):
        s = atm + k * step
        match = next((r for r in chain if r.get('strike') == s), None)
        if match:
            strikes_obj[s] = {
                'ce_iv':  match.get('ce_iv', 0)  or 0,
                'pe_iv':  match.get('pe_iv', 0)  or 0,
                'ce_ltp': match.get('ce_ltp', 0) or 0,
                'pe_ltp': match.get('pe_ltp', 0) or 0,
            }

    ist = timezone(timedelta(hours=5, minutes=30))
    now = datetime.now(ist)
    time_str = now.strftime('%H:%M')

    # Get current VIX from market overview.
    # _market_cache is populated by refresh_market_overview() via _market_cache.update(result),
    # so VIX lives at _market_cache['VIX'], not _market_cache['overview']['VIX'].
    # Fallback: if market overview hasn't populated yet (first few minutes after startup),
    # pull VIX from the latest iv_history entry which is captured on the same poll cycle.
    vix_val = 0
    try:
        vix_val = (_market_cache.get('VIX') or {}).get('last', 0) or 0
    except Exception:
        vix_val = 0
    if not vix_val:
        try:
            iv_hist = get_iv_history(symbol) or []
            if iv_hist:
                vix_val = iv_hist[-1].get('vix', 0) or 0
        except Exception:
            pass

    snap = {
        'time':    time_str,
        'spot':    spot,
        'vix':     vix_val,
        'atm':     atm,
        'step':    step,
        'strikes': strikes_obj,
    }

    history = _strike_iv_history.setdefault(symbol, [])

    # Backfill vix=0 in any historical rows now that we have a valid VIX.
    # This heals snapshots captured before refresh_market_overview populated _market_cache.
    if vix_val:
        for h in history:
            if not h.get('vix'):
                h['vix'] = vix_val

    # Avoid duplicates within the same minute
    if history and history[-1]['time'] == time_str:
        history[-1] = snap   # update in place
    else:
        history.append(snap)
        # Cap at 130 snapshots (~6.5 hours of 3-min polls = full session)
        if len(history) > 130:
            history.pop(0)


def refresh_options():
    symbols = ['NIFTY', 'BANKNIFTY']
    while True:
        print("\n--- Refreshing options ---")
        for symbol in symbols:
            try:
                result = fetch_options(symbol)
                if result:
                    _options_cache[symbol] = result
                    print(f"  [options] {symbol} done — PCR: {result.get('pcr_total')}")
                    # Snapshot ATM±5 strike IVs into rolling history (used by frontend picker + EOD export)
                    try:
                        chain_data = get_full_chain(symbol)
                        if chain_data:
                            _snapshot_strike_iv(symbol, chain_data)
                    except Exception as e:
                        print(f"  [strike_iv] {symbol} snapshot failed: {e}")
                    try:
                        fut_payload = get_latest(symbol)
                        prev_pcr    = _prev_pcr_3strike.get(symbol)
                        gamma       = compute_gamma_blast(result, fut_payload, prev_pcr)
                        if gamma:
                            _gamma_cache[symbol]      = gamma
                            _prev_pcr_3strike[symbol] = result.get('pcr_3strike', 0)
                            print(f"  [gamma] {symbol} rating:{gamma['rating']} score:{gamma['score']} dir:{gamma['direction']}")
                    except Exception as e:
                        print(f"  [gamma] {symbol} failed: {e}")
            except Exception as e:
                print(f"  [options] {symbol} failed: {e}")
        time.sleep(180)

def refresh_indices():
    while True:
        print("\n--- Refreshing indices ---")
        try:
            result  = indices.run()
            cache.update_indices(result)
            payload = safe_json({'type': 'indices', 'data': result})
            cache.broadcast(payload)
            print(f"--- Indices done: {len(result)} ---")
        except Exception as e:
            print(f"  [indices] failed: {e}")
        time.sleep(60)

def refresh_stocks():
    from screeners.base import fetch_daily, base_stock_info, get_symbols
    from screeners.nse_stocks import get_nifty50_stocks
    while True:
        for country in COUNTRIES:
            print(f"\n--- Refreshing stocks [{country}] ---")
            if country == 'IN':
                try:
                    results = get_nifty50_stocks()
                    cache.update_stocks('IN', results)
                    print(f"--- NSE stocks done: {len(results)} ---")
                except Exception as e:
                    print(f"  [stocks] NSE error: {e}")
            else:
                results = []
                for symbol in get_symbols(country):
                    try:
                        hist, info = fetch_daily(symbol)
                        if hist is not None and len(hist) >= 2:
                            base = base_stock_info(symbol, hist, info)
                            if base:
                                base['country'] = country
                                results.append(base)
                    except Exception as e:
                        print(f"  [stocks] error {symbol}: {e}")
                cache.update_stocks(country, results)
                print(f"--- Stocks done [{country}]: {len(results)} ---")
        time.sleep(180)

def refresh_loop():
    while True:
        for country in COUNTRIES:
            print(f"\n--- Running screeners [{country}] ---")
            all_results = {}
            for screener in SCREENERS:
                try:
                    print(f"  Running {screener.SCREENER_ID}...")
                    results = screener.run(country=country)
                    all_results[screener.SCREENER_ID] = results
                except Exception as e:
                    print(f"  [app] {screener.SCREENER_ID} failed: {e}")
                    all_results[screener.SCREENER_ID] = []
            cache.update(country, all_results)
            payload = safe_json({
                'type':    'update',
                'country': country,
                'data':    all_results,
            })
            cache.broadcast(payload)
            print(f"--- Screeners done [{country}] ---")
        print("Sleeping 60s...")
        time.sleep(60)

@app.route('/api/indices')
def get_indices():
    return jsonify(sanitize(cache.get_indices()))

@app.route('/api/stocks')
def get_stocks():
    country = request.args.get('country', 'PL')
    return jsonify(sanitize(cache.get_stocks(country)))

@app.route('/api/stocks/wig30')
def get_wig30():
    return jsonify(sanitize(cache.get_stocks('PL')))

@app.route('/api/screeners')
def get_screeners():
    country = request.args.get('country', 'PL')
    return jsonify(sanitize(cache.get(country)))

@app.route('/api/sector-scope')
def get_sector_scope():
    country = request.args.get('country', 'PL')
    data    = cache.get(country)
    return jsonify(sanitize(data.get('sector_scope', {})))

@app.route('/api/sessions')
def get_sessions():
    return jsonify(sanitize(get_all_sessions()))

@app.route('/api/market-overview')
def get_market_overview_route():
    if not _market_cache:
        try:
            result = get_market_overview()
            _market_cache.update(result)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    return jsonify(sanitize(_market_cache))

@app.route('/api/options')
def get_options():
    symbol = request.args.get('symbol', 'NIFTY').upper()
    if symbol not in ['NIFTY', 'BANKNIFTY']:
        return jsonify({'error': 'Invalid symbol'}), 400
    data = _options_cache.get(symbol)
    if not data:
        try:
            data = fetch_options(symbol)
            if data:
                _options_cache[symbol] = data
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    return jsonify(sanitize(data or {}))

@app.route('/api/option-chain')
def get_option_chain_full():
    symbol = request.args.get('symbol', 'NIFTY').upper()
    if symbol not in ['NIFTY', 'BANKNIFTY']:
        return jsonify({'error': 'Invalid symbol'}), 400
    data = get_full_chain(symbol)
    if not data:
        try:
            result = fetch_options(symbol)
            if result:
                _options_cache[symbol] = result
                data = get_full_chain(symbol)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    return jsonify(sanitize(data or {}))

@app.route('/api/volume/analysis')
def get_volume_analysis():
    """
    Returns enriched per-strike volume analytics for ATM ±10 strikes:
      - V/OI ratios + buildup classification per side
      - CE/PE volume split, churn %, dominant buildup
      - Top 5 strikes by volume with smart-money score

    Computed on-demand from current _options_cache and the previous snapshot
    (held in _prev_chain_cache) to derive delta volume / delta OI per snapshot.
    """
    symbol = request.args.get('symbol', 'NIFTY').upper()
    if symbol not in ['NIFTY', 'BANKNIFTY']:
        return jsonify({'error': 'Invalid symbol'}), 400

    data = get_full_chain(symbol)
    if not data:
        try:
            result = fetch_options(symbol)
            if result:
                _options_cache[symbol] = result
                data = get_full_chain(symbol)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    if not data:
        return jsonify({'error': 'No chain data available'}), 503

    prev = _prev_chain_cache.get(symbol)

    try:
        analysis = compute_volume_analytics(data, prev)
    except Exception as e:
        print(f"  [volume] {symbol} compute failed: {e}")
        return jsonify({'error': str(e)}), 500

    # Update prev cache AFTER computation so the next call uses this snapshot as prev
    _prev_chain_cache[symbol] = data

    return jsonify(sanitize(analysis or {}))

@app.route('/api/strike-iv-history')
def get_strike_iv_history_route():
    """
    Returns rolling per-strike IV history for ATM±5 strikes.
    Captured every 3 min by refresh_options() — same cadence as the option chain.
    Used by the IV Spread + Demand table's strike picker on the frontend.
    """
    symbol = request.args.get('symbol', 'NIFTY').upper()
    if symbol not in ['NIFTY', 'BANKNIFTY']:
        return jsonify({'error': 'Invalid symbol'}), 400
    history = _strike_iv_history.get(symbol, [])
    return jsonify(sanitize({'history': history}))

@app.route('/api/gamma-blast')
def get_gamma_blast():
    symbol = request.args.get('symbol', 'NIFTY').upper()
    if symbol not in ['NIFTY', 'BANKNIFTY']:
        return jsonify({'error': 'Invalid symbol'}), 400
    data = _gamma_cache.get(symbol)
    if not data:
        options = _options_cache.get(symbol)
        if options:
            try:
                fut  = get_latest(symbol)
                data = compute_gamma_blast(options, fut, _prev_pcr_3strike.get(symbol))
                if data:
                    _gamma_cache[symbol] = data
            except Exception as e:
                return jsonify({'error': str(e)}), 500
    return jsonify(sanitize(data or {}))

@app.route('/api/futures-sentiment')
def get_futures_sentiment():
    symbol = request.args.get('symbol', 'NIFTY').upper()
    if symbol not in ['NIFTY', 'BANKNIFTY']:
        return jsonify({'error': 'Invalid symbol'}), 400
    data = get_latest(symbol)
    if not data:
        try:
            data = poll_futures_sentiment(symbol)
            if data:
                update_latest(symbol, data)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    return jsonify(sanitize(data or {}))

@app.route('/api/options/pcr-history')
def get_pcr_history_route():
    symbol = request.args.get('symbol', 'NIFTY').upper()
    return jsonify(sanitize(get_pcr_history(symbol)))

@app.route('/api/futures-dashboard')
def get_futures_dashboard():
    symbol = request.args.get('symbol', 'NIFTY').upper()
    if symbol not in ['NIFTY', 'BANKNIFTY']:
        return jsonify({'error': 'Invalid symbol'}), 400
    options_data = _options_cache.get(symbol)
    result = fetch_futures_dashboard(symbol, options_data)
    return jsonify(sanitize(result or {}))

@app.route('/api/export-eod')
def trigger_eod_export():
    try:
        path = export_eod_spreadsheet()
        return jsonify({'status': 'ok', 'file': os.path.basename(path)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/countries')
def get_countries():
    return jsonify([
        {'code': code, 'label': COUNTRY_LABELS[code]}
        for code in COUNTRIES
    ])

@sock.route('/ws')
def websocket(ws):
    cache.add_client(ws)
    print(f"Client connected")
    all_data = cache.get()
    for country, data in all_data.items():
        if data:
            ws.send(safe_json({
                'type':    'update',
                'country': country,
                'data':    data,
            }))
    idx_data = cache.get_indices()
    if idx_data:
        ws.send(safe_json({'type': 'indices', 'data': idx_data}))
    try:
        while True:
            msg = ws.receive(timeout=30)
            if msg is None:
                break
    except Exception:
        pass
    finally:
        cache.remove_client(ws)
        print(f"Client disconnected")

if __name__ == '__main__':
    t1 = threading.Thread(target=refresh_stocks,          daemon=True)
    t1.start()
    t2 = threading.Thread(target=refresh_loop,            daemon=True)
    t2.start()
    t3 = threading.Thread(target=refresh_indices,         daemon=True)
    t3.start()
    t4 = threading.Thread(target=refresh_options,         daemon=True)
    t4.start()
    t5 = threading.Thread(target=refresh_market_overview, daemon=True)
    t5.start()
    t_futures = threading.Thread(target=refresh_futures,  daemon=True)
    t_futures.start()
    t_news = threading.Thread(target=refresh_news,        daemon=True)
    t_news.start()
    t_fd = threading.Thread(target=refresh_futures_dashboard, daemon=True)
    t_fd.start()
    t_eod = threading.Thread(target=eod_scheduler, daemon=True)
    t_eod.start()
    t_candles = threading.Thread(target=refresh_candles, daemon=True)
    t_candles.start()
    app.run(port=3001, debug=False)